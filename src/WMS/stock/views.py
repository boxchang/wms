import json

import openpyxl
from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.views.decorators.csrf import csrf_exempt
from suds import client
from django.http import JsonResponse
from WMS.settings.base import HANEL_WS
from stock.forms import SearchForm
from stock.models import Storage_Type, Storage, Location, Bin, Item, WO


@login_required
def wo_search(request):
    if request.method == 'POST':
        wo_no = request.POST.get('wo_no')
        items = WO.objects.filter(wo_no=wo_no).all().order_by('item__item_code')
        search_form = SearchForm(initial={'wo_no': wo_no})
    else:
        search_form = SearchForm()
    return render(request, 'stock/wo_search.html', locals())


@login_required
def wo_import(request):
    if request.method == 'POST':
        excel_file = request.FILES.get('files1')
        wo_del_list = []
        if excel_file:
            wb = openpyxl.load_workbook(excel_file)
            sheet = wb.worksheets[0]
            for iRow in range(2, sheet.max_row + 1):
                try:
                    wo_no = sheet.cell(row=iRow, column=1).value
                    wo = WO.objects.filter(wo_no=wo_no)
                    if wo and wo_no not in wo_del_list:
                        wo.delete()
                        wo_del_list.append(wo_no)

                    item_code = str(sheet.cell(row=iRow, column=2).value)
                    item = Item.objects.get(item_code=item_code)
                    qty = sheet.cell(row=iRow, column=3).value
                    wo = WO.objects.create(wo_no=wo_no, item=item,
                                           qty=qty, update_by=request.user)

                except Exception as e:
                    print(e)

    return render(request, 'stock/wo_import.html', locals())


@login_required
def item_import(request):
    if request.method == 'POST':
        excel_file = request.FILES.get('files1')
        if excel_file:
            wb = openpyxl.load_workbook(excel_file)
            sheet = wb.worksheets[0]
            for iRow in range(2, sheet.max_row + 1):
                try:
                    item_code = str(sheet.cell(row=iRow, column=1).value)
                    if len(item_code) == 10:
                        desc = sheet.cell(row=iRow, column=2).value
                        new_bin = sheet.cell(row=iRow, column=7).value
                        new_bin = Bin.objects.get(bin_code=new_bin)
                        item = Item.objects.update_or_create(item_code=item_code, defaults={'desc': desc, 'update_by': request.user})
                        item = Item.objects.get(item_code=item_code)
                        bins = item.bin.all()
                        if new_bin not in bins:
                            item.bin.add(new_bin)
                except Exception as e:
                    print(e)

    return render(request, 'stock/import.html', locals())

@login_required
def excel_import(request):
    if request.method == 'POST':
        excel_file = request.FILES.get('files1')
        if excel_file:
            try:
                wb = openpyxl.load_workbook(excel_file)
                for sheet in wb.worksheets:
                    print(sheet)
                    # if sheet.title == "Storage":
                    #     for iRow in range(2, sheet.max_row + 1):
                    #         storage_code = sheet.cell(row=iRow, column=1).value
                    #         type = sheet.cell(row=iRow, column=2).value
                    #         type = Storage_Type.objects.get(type_code=type)
                    #         desc = sheet.cell(row=iRow, column=3).value
                    #         ip_addr = sheet.cell(row=iRow, column=4).value
                    #         lift = sheet.cell(row=iRow, column=5).value
                    #         access_point = sheet.cell(row=iRow, column=6).value
                    #         enable = sheet.cell(row=iRow, column=7).value
                    #         storage = Storage.objects.update_or_create(storage_code=storage_code,
                    #                                                    defaults={'type': type, 'desc': desc,
                    #                                                              'ip_addr': ip_addr,
                    #                                                              'lift': lift,
                    #                                                              'access_point': access_point,
                    #                                                              'update_by': request.user,
                    #                                                              'enable': enable, })
                    if sheet.title == "Location":
                        for iRow in range(2, sheet.max_row + 1):
                            storage = sheet.cell(row=iRow, column=1).value
                            storage = Storage.objects.get(storage_code=storage)
                            location_code = sheet.cell(row=iRow, column=2).value
                            location_name = sheet.cell(row=iRow, column=3).value
                            mach_location_code = sheet.cell(row=iRow, column=4).value
                            type = sheet.cell(row=iRow, column=5).value
                            type = Storage_Type.objects.get(type_code=type)
                            desc = sheet.cell(row=iRow, column=6).value
                            enable = sheet.cell(row=iRow, column=7).value
                            location = Location.objects.update_or_create(storage=storage, location_code=location_code,
                                                                 defaults={'location_name':location_name, 'mach_location_code':mach_location_code,
                                                                         'type':type, 'desc':desc, 'enable':enable, 'update_by':request.user})
                    elif sheet.title == "Bin":
                        for iRow in range(2, sheet.max_row + 1):
                            bin_code = sheet.cell(row=iRow, column=1).value
                            bin_name = sheet.cell(row=iRow, column=2).value
                            location_code = sheet.cell(row=iRow, column=4).value
                            location = Location.objects.get(location_code=location_code)
                            desc = sheet.cell(row=iRow, column=5).value
                            enable = sheet.cell(row=iRow, column=6).value
                            bin = Bin.objects.update_or_create(location=location, bin_code=bin_code,
                                                                 defaults={'bin_name': bin_name, 'enable': enable,
                                                                    'update_by': request.user})
            except Exception as e:
                print(e)


    return render(request, 'stock/import.html', locals())



@login_required
def search(request):
    storages = Storage.objects.all()
    return render(request, 'stock/storage.html', locals())


@login_required
def location_list(request):
    if request.method == 'POST':
        storage_code = request.POST.get('pk')
        storage = Storage.objects.get(storage_code=storage_code)
        locations = Location.objects.filter(storage=storage).all()
    return render(request, 'stock/location.html', locals())


@login_required
def bin_list(request, storage, location):
    bins = Bin.objects.filter(location=location).all()
    return render(request, 'stock/bin.html', locals())


@csrf_exempt
def call_location(request):
    if request.method == 'POST':
        try:
            storage = request.POST.get('storage')
            location = request.POST.get('location')
            lift = request.POST.get('lift')
            access_point = request.POST.get('access_point')

            web_s = client.Client(HANEL_WS)
            params = web_s.factory.create("ns2:JobRecordV02")
            params.lift = lift
            params.accessPoint = access_point  # 1:一樓   2:二樓
            params.shelf = location  # 01;02;03
            params.job = '1000727161'
            res = web_s.service.sendJobBufferV02(params)
            result = "DONE"
        except Exception as e:
            result = "ERROR"

    return JsonResponse(result, safe=False)


@csrf_exempt
def call_wo_item_check(request):
    if request.method == 'POST':
        try:
            pk = request.POST.get('pk')
            item_checked = request.POST.get('item_checked')
            wo_item = WO.objects.get(pk=pk)
            wo_item.checked = json.loads(item_checked)
            wo_item.save()
            result = "DONE"
        except Exception as e:
            result = "ERROR"

    return JsonResponse(result, safe=False)